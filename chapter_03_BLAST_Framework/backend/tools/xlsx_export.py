"""Layer 3 Tool: write generated test plans to a single .xlsx deliverable.

Deterministic and atomic. Matches the "Final on-disk artifact" contract in
claude.md: one .xlsx with columns Test Case ID, Title, Type, Priority,
Preconditions, Steps, Expected Result, Story ID, AC Ref.
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_COLUMNS = [
    "Test Case ID", "Title", "Type", "Priority", "Preconditions",
    "Steps", "Expected Result", "Story ID", "AC Ref",
]


def export_test_plans(test_plans: list[dict], output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Plan"

    ws.append(_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")

    for plan in test_plans:
        for tc in plan["test_cases"]:
            ws.append([
                tc["test_case_id"],
                tc["title"],
                tc["type"],
                tc["priority"],
                tc["preconditions"],
                " | ".join(tc["steps"]),
                tc["expected_result"],
                tc["traceability"]["story_id"],
                tc["traceability"]["acceptance_criteria_ref"],
            ])

    for i, column_title in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(60, len(column_title) + 10))

    output_dir.mkdir(exist_ok=True)
    story_ids = "_".join(str(p["story_id"]) for p in test_plans) or "empty"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"test_plan_{story_ids}_{timestamp}.xlsx"
    path = output_dir / filename
    wb.save(path)
    return path
